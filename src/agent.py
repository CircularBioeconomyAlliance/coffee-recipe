#!/usr/bin/env python3
"""Simple Strands chatbot with Bedrock knowledge base support."""

import argparse
from pathlib import Path

import openpyxl
from pypdf import PdfReader
from strands import Agent
from strands.models import BedrockModel
from strands_tools import current_time, file_write, http_request, memory, use_llm


def extract_pdf_text(pdf_path: str) -> str:
    """Extract text from a PDF file."""
    reader = PdfReader(pdf_path)
    text = []
    for page in reader.pages:
        text.append(page.extract_text())
    return "\n".join(text)


def extract_excel_text(excel_path: str) -> str:
    """Extract text from an Excel file."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text.append(f"=== Sheet: {sheet_name} ===")
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell) if cell is not None else "" for cell in row]
            if any(row_text):  # Skip empty rows
                text.append(" | ".join(row_text))
    return "\n".join(text)


def extract_file_text(file_path: Path) -> str:
    """Extract text from PDF or Excel file."""
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf_text(str(file_path))
    elif suffix in [".xlsx", ".xls"]:
        return extract_excel_text(str(file_path))
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Use .pdf or .xlsx")


# Import shared configuration
from config import MODEL_ID, KNOWLEDGE_BASE_ID, SYSTEM_PROMPT


def main():
    parser = argparse.ArgumentParser(description="CBA Indicator Selection Agent")
    parser.add_argument(
        "--file", type=str, help="Path to PDF or Excel file with project description"
    )
    args = parser.parse_args()

    print(f"Using Knowledge Base ID: {KNOWLEDGE_BASE_ID}")

    # Configure model with low temperature for deterministic responses
    bedrock_model = BedrockModel(
        model_id=MODEL_ID,
        temperature=0.2,  # Low temperature = less creative, more focused
    )

    agent = Agent(
        model=bedrock_model,
        system_prompt=SYSTEM_PROMPT,
        tools=[
            memory,
            use_llm,
            http_request,
            file_write,
            current_time,
        ],
    )

    print("\nCBA Knowledge Base Chatbot")
    print("Type 'exit' to quit\n")

    # If file provided, extract text and send as first message
    if args.file:
        file_path = Path(args.file)
        if file_path.exists():
            print(f"Reading file: {file_path}")
            try:
                file_text = extract_file_text(file_path)
                initial_prompt = f"Here is my project description:\n\n{file_text}\n\nPlease recommend indicators for this project."
                print(f"> [Content from {file_path.name}]\n")
                agent(initial_prompt)
                print()
            except ValueError as e:
                print(f"Error: {e}")
                return
        else:
            print(f"Error: File not found: {file_path}")
            return

    while True:
        user_input = input("> ")

        if user_input.lower() in ["exit", "quit"]:
            print("\nGoodbye!")
            break

        if not user_input.strip():
            continue

        agent(user_input)
        print()


if __name__ == "__main__":
    main()
